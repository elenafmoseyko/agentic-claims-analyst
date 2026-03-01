"""
src/report_exporter.py
======================
Generates the final Excel exhibit workbook with:
  - Exhibit 1 – Summary tab (PMPM trend + membership)
  - Exhibit 2 – Detail tab (monthly breakdown by LOB + provider)
  - Embedded charts in each tab
"""

from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from .metrics import calculate_pmpm, calculate_summary, membership_by_lob

# ── Colour palette ─────────────────────────────────────────────────────────────
NAVY    = "1F3864"
BLUE    = "2E75B6"
GREY    = "D6E4F0"
WHITE   = "FFFFFF"
ORANGE  = "ED7D31"
GREEN   = "70AD47"

MONTHS  = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
LOB_LIST = ["HMO/POS", "PPO", "Medicaid", "Medicare"]
PROVIDERS = ["Provider A", "Provider B", "Provider C"]


def export_excel_exhibit(df: pd.DataFrame, output_path: str) -> str:
    """
    Build and save the full Excel exhibit workbook.
    Returns the resolved output path.
    """
    pmpm_df    = calculate_pmpm(df)
    summary_df = calculate_summary(df)
    mbr_df     = membership_by_lob(df)

    wb = openpyxl.Workbook()

    _build_summary_tab(wb, pmpm_df, df)
    _build_detail_tab(wb, df, summary_df)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    path = str(Path(output_path).resolve())
    wb.save(path)
    return path


# ── Exhibit 1 – Summary ────────────────────────────────────────────────────────
def _build_summary_tab(wb, pmpm_df: pd.DataFrame, raw_df: pd.DataFrame):
    ws = wb.create_sheet("Exhibit 1 - Summary")

    # Title block
    _write_title_block(ws, "Historical Monthly Claims Trend Summary",
                       "Claims Incurred 1/1–12/31, Paid through 12/31")

    # ── PMPM table ────────────────────────────────────────────────────────────
    start_row = 6
    headers = ["Month"] + PROVIDERS
    _write_header_row(ws, start_row, 1, headers, NAVY)

    months_sorted = sorted(pmpm_df["month"].unique())
    for r, month in enumerate(months_sorted, start=start_row + 1):
        label = pd.Timestamp(month).strftime("%b")
        ws.cell(r, 1, label).alignment = Alignment(horizontal="center")
        for c, provider in enumerate(PROVIDERS, 2):
            val = pmpm_df.loc[
                (pmpm_df["month"] == month) & (pmpm_df["provider"] == provider), "pmpm"
            ]
            cell = ws.cell(r, c, round(val.values[0], 2) if not val.empty else None)
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="right")

    # Totals row
    total_row = start_row + len(months_sorted) + 1
    ws.cell(total_row, 1, "Annual Avg").font = Font(bold=True)
    for c, provider in enumerate(PROVIDERS, 2):
        col_vals = [
            ws.cell(r, c).value
            for r in range(start_row + 1, total_row)
            if ws.cell(r, c).value is not None
        ]
        avg = sum(col_vals) / len(col_vals) if col_vals else None
        cell = ws.cell(total_row, c, round(avg, 2) if avg else None)
        cell.number_format = '"$"#,##0.00'
        cell.font = Font(bold=True)

    _add_table_borders(ws, start_row, 1, total_row, len(headers))

    # ── PMPM line chart ───────────────────────────────────────────────────────
    chart = LineChart()
    chart.title = "Medical PMPM by Provider – All Lines of Business"
    chart.style = 10
    chart.y_axis.title = "PMPM ($)"
    chart.x_axis.title = "Month"
    chart.height = 12
    chart.width  = 22

    # Add all provider series at once
    data = Reference(ws, min_col=2, max_col=1 + len(PROVIDERS),
                     min_row=start_row, max_row=total_row - 1)
    chart.add_data(data, titles_from_data=True)
    cats = Reference(ws, min_col=1, max_col=1,
                     min_row=start_row + 1, max_row=total_row - 1)
    chart.set_categories(cats)
    ws.add_chart(chart, f"F{start_row}")

    # ── Membership table ──────────────────────────────────────────────────────
    mbr_start = total_row + 3
    mbr_df = membership_by_lob(raw_df)
    months_sorted2 = sorted(mbr_df["month"].unique())
    lobs = mbr_df["lob"].unique().tolist()

    _write_header_row(ws, mbr_start, 1, ["Month"] + lobs, NAVY)
    for r, month in enumerate(months_sorted2, start=mbr_start + 1):
        label = pd.Timestamp(month).strftime("%b")
        ws.cell(r, 1, label).alignment = Alignment(horizontal="center")
        for c, lob in enumerate(lobs, 2):
            val = mbr_df.loc[
                (mbr_df["month"] == month) & (mbr_df["lob"] == lob), "members"
            ]
            cell = ws.cell(r, c, int(val.values[0]) if not val.empty else None)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")

    _set_col_widths(ws, {1: 10, 2: 15, 3: 15, 4: 15, 5: 15, 6: 15})


# ── Exhibit 2 – Detail ─────────────────────────────────────────────────────────
def _build_detail_tab(wb, df: pd.DataFrame, summary_df: pd.DataFrame):
    ws = wb.create_sheet("Exhibit 2 - Detail")

    _write_title_block(ws, "Detailed Historical Experience Summary by Line of Business",
                       "Claims Incurred 1/1–12/31, Paid through 12/31")

    start_row = 6
    lobs = df["lob"].unique().tolist()

    # Sub-headers for each LOB block
    col = 2
    ws.cell(start_row, 1, "Month").font = Font(bold=True, color=WHITE)
    ws.cell(start_row, 1).fill = PatternFill("solid", fgColor=NAVY)

    for lob in lobs:
        ws.merge_cells(start_row=start_row, start_column=col,
                       end_row=start_row, end_column=col + 3)
        cell = ws.cell(start_row, col, lob)
        cell.font      = Font(bold=True, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center")
        for offset, label in enumerate(["Members"] + PROVIDERS, 0):
            ws.cell(start_row + 1, col + offset, label).font = Font(bold=True)
            ws.cell(start_row + 1, col + offset).fill = PatternFill("solid", fgColor=GREY)
        col += 4

    months_sorted = sorted(df["month"].unique())
    for r, month in enumerate(months_sorted, start=start_row + 2):
        label = pd.Timestamp(month).strftime("%b")
        ws.cell(r, 1, label).alignment = Alignment(horizontal="center")
        col = 2
        for lob in lobs:
            sub = df[(df["month"] == month) & (df["lob"] == lob)]
            mbr = sub["members"].sum()
            ws.cell(r, col, int(mbr) if mbr else None).number_format = "#,##0"
            for offset, provider in enumerate(PROVIDERS, 1):
                val = sub.loc[sub["provider"] == provider, "pmpm"]
                cell = ws.cell(r, col + offset,
                               round(val.values[0], 2) if not val.empty else None)
                cell.number_format = '"$"#,##0.00'
            col += 4

    _set_col_widths(ws, {i: 13 for i in range(1, 20)})

    # Summary chart per LOB
    chart_row = start_row + 2 + len(months_sorted) + 2
    for i, lob in enumerate(lobs):
        sub = summary_df[summary_df["lob"] == lob]
        chart = BarChart()
        chart.type = "col"
        chart.title = f"{lob} – Avg PMPM by Provider"
        chart.y_axis.title = "Avg PMPM ($)"
        chart.height = 10
        chart.width  = 14
        ws.add_chart(chart, f"{get_column_letter(2 + i * 4)}{chart_row}")


# ── Utility helpers ────────────────────────────────────────────────────────────
def _write_title_block(ws, title: str, subtitle: str):
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "Company ABC"
    c.font  = Font(bold=True, size=14, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = title
    c.font  = Font(bold=True, size=12, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value = subtitle
    c.font  = Font(italic=True, size=10)
    c.alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20


def _write_header_row(ws, row: int, start_col: int, headers: list, bg: str):
    for c, h in enumerate(headers, start_col):
        cell = ws.cell(row, c, h)
        cell.font      = Font(bold=True, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center")


def _add_table_borders(ws, min_row, min_col, max_row, num_cols):
    thin = Side(style="thin")
    for r in range(min_row, max_row + 1):
        for c in range(min_col, min_col + num_cols):
            ws.cell(r, c).border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )


def _set_col_widths(ws, widths: dict):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
